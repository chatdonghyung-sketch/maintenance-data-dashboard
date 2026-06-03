"""TREND_20260331.xls 에 데모 알람 패턴 주입.
임계초과 ×2 / 도달예측 ×2 / SPC ×2 가 알람 감지에 뜨도록 6개 태그의 24시간 값을 덮어쓴다.
알람값(HA/LA)·메타는 그대로 두고 시간값만 수정하므로, 값이 달라져도 패턴은 상대적으로 생성된다.

사용법:  Excel에서 파일을 닫은 뒤  ->  python inject_demo_alarms.py
"""
import openpyxl, random
from io import BytesIO

PATH = 'HMI_Trend_data/TREND_20260331.xls'
HCOL = range(7, 31)  # 24시간 컬럼 (1-indexed 7..30)


def main():
    with open(PATH, 'rb') as f:
        wb = openpyxl.load_workbook(BytesIO(f.read()))
    ws = wb.active

    def info(r):
        return dict(tag=ws.cell(r, 5).value, HA=ws.cell(r, 31).value,
                    LA=ws.cell(r, 32).value, unit=ws.cell(r, 33).value)

    def setrow(r, vals):
        for c, v in zip(HCOL, vals):
            ws.cell(r, c).value = round(v, 3)

    def inject_violation(r, HA, LA):
        rng = HA - LA; mid = (HA + LA) / 2
        mx = []; av = []; mn = []
        for i in range(24):
            a = mid + random.uniform(-0.05, 0.05) * rng
            m = a + 0.15 * rng
            if i in (5, 9, 13, 17, 21):
                m = HA + 0.12 * rng        # 상한 초과 시간
            mx.append(m); av.append(a); mn.append(a - 0.15 * rng)
        setrow(r, mx); setrow(r + 1, av); setrow(r + 2, mn)

    def inject_forecast(r, HA, LA):
        rng = HA - LA; mid = (HA + LA) / 2
        s = mid; e = HA - 0.04 * rng       # 상한 직전까지 깨끗이 상승
        mx = []; av = []; mn = []
        for i in range(24):
            m = s + (e - s) * (i / 23) + random.uniform(-0.01, 0.01) * rng
            mx.append(min(m, HA - 0.02 * rng)); av.append(m - 0.08 * rng); mn.append(m - 0.16 * rng)
        setrow(r, mx); setrow(r + 1, av); setrow(r + 2, mn)

    def inject_spc(r, HA, LA):
        rng = HA - LA; mid = (HA + LA) / 2
        lo = mid - 0.25 * rng; hi = mid + 0.25 * rng   # 후반 14h부터 평균 이동
        mx = []; av = []; mn = []
        for i in range(24):
            a = (lo if i < 14 else hi) + random.uniform(-0.02, 0.02) * rng
            mx.append(a + 0.08 * rng); av.append(a); mn.append(a - 0.08 * rng)
        setrow(r, mx); setrow(r + 1, av); setrow(r + 2, mn)

    random.seed(331)
    plan = [
        (3,  inject_violation), (6,  inject_violation),   # 임계초과 ×2
        (9,  inject_forecast),  (12, inject_forecast),    # 도달예측 ×2
        (15, inject_spc),       (18, inject_spc),         # SPC ×2
    ]
    applied = []
    for r, fn in plan:
        d = info(r)
        fn(r, float(d['HA']), float(d['LA']))
        applied.append((d['tag'], fn.__name__.replace('inject_', ''), d['HA'], d['LA'], d['unit']))

    wb.save(PATH)
    print('주입 완료:')
    for t, k, ha, la, u in applied:
        print(f'  {t:12s} {k:10s} HA={ha} LA={la} {u}')


if __name__ == '__main__':
    main()
