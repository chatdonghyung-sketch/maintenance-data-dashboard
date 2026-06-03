from contextlib import asynccontextmanager
from fastapi import FastAPI, Query, Request, UploadFile, File
from fastapi.responses import JSONResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from typing import List
import os
import glob
import re
import random
import xlrd
import openpyxl
import io

from database import (
    init_db,
    load_chiller_month, save_chiller_cell, save_chiller_inspector, save_chiller_bulk,
    load_chiller_analysis, save_chiller_analysis_cell, save_chiller_analysis_bulk,
    load_date_form, save_date_form_cell, save_date_form_bulk,
    upsert_work_orders, query_work_orders, get_work_order_kpi,
    get_work_order_filter_options, upsert_work_order_single, update_work_order_status,
    upsert_energy_usage, query_energy_db, delete_energy_usage, delete_energy_usage_range,
    get_cost_records, upsert_cost_record, delete_cost_record,
    get_inventory_items, add_inventory_item, update_inventory_qty,
    add_inventory_issue, get_inventory_issues,
    get_compliance_officers, add_compliance_officer, delete_compliance_officer,
    get_compliance_educations, add_compliance_education, delete_compliance_education,
    get_compliance_inspections, add_compliance_inspection, delete_compliance_inspection,
    get_preventive_plans, add_preventive_plan, update_preventive_plan_status,
    get_preventive_results, add_preventive_result,
    get_checksheets, add_checksheet, add_checksheet_item,
    add_alarm_action, get_alarm_actions,
    get_dashboard_summary,
)


@asynccontextmanager
async def lifespan(app: FastAPI):
    await init_db()
    # 서버 시작 시 전공장 엑셀 파일 자동 가져오기 (파일 없으면 무시)
    try:
        all_records = []
        for cfg in ZANGONG_FILES:
            path = os.path.join(PROJECT_ROOT, cfg['folder'], cfg['filename'])
            if not os.path.exists(path):
                continue
            wb = openpyxl.load_workbook(path, data_only=True)
            records = (_parse_zangong_horizontal(wb, cfg['util_key'])
                       if cfg['format'] == 'horizontal'
                       else _parse_zangong_vertical(wb, cfg['util_key']))
            wb.close()
            all_records.extend(records)
        if all_records:
            await upsert_energy_usage(all_records)
    except Exception:
        pass
    yield


app = FastAPI(lifespan=lifespan)

TREND_DIR = os.path.normpath(os.path.join(os.path.dirname(__file__), '..', 'HMI_Trend_data'))
DIST_DIR = os.path.normpath(os.path.join(os.path.dirname(__file__), '..', 'dist'))


# ── 냉동기·냉각탑 운전일지 ────────────────────────────────────────────
@app.get('/api/chiller-log')
async def get_chiller_log(year: int = Query(...), month: int = Query(...)):
    data = await load_chiller_month(year, month)
    return JSONResponse(data)


@app.post('/api/chiller-log/cell')
async def save_chiller_cell_route(request: Request):
    body = await request.json()
    if body.get('type') == 'inspector':
        await save_chiller_inspector(body['year'], body['month'], body['day'], body['value'])
    else:
        await save_chiller_cell(body['year'], body['month'], body['day'],
                                body['equipment_id'], body['param_key'], body['value'])
    return JSONResponse({'ok': True})


@app.post('/api/chiller-log/bulk')
async def save_chiller_bulk_route(request: Request):
    body = await request.json()
    await save_chiller_bulk(body['year'], body['month'],
                            body.get('cells', []), body.get('inspectors', []))
    return JSONResponse({'ok': True})


# ── 냉동기 가동 분석 ──────────────────────────────────────────────────
@app.get('/api/chiller-analysis')
async def get_chiller_analysis(year: int = Query(...), month: int = Query(...)):
    data = await load_chiller_analysis(year, month)
    return JSONResponse(data)


@app.post('/api/chiller-analysis/cell')
async def save_analysis_cell(request: Request):
    body = await request.json()
    await save_chiller_analysis_cell(body['year'], body['month'], body['day'],
                                      body['chiller_id'], body['value'])
    return JSONResponse({'ok': True})


@app.post('/api/chiller-analysis/bulk')
async def save_analysis_bulk(request: Request):
    body = await request.json()
    await save_chiller_analysis_bulk(body['year'], body['month'], body.get('records', []))
    return JSONResponse({'ok': True})


# ── 일별 폼 공용 라우트 (FFU / Boiler / Safety) ───────────────────────
@app.get('/api/ffu-log')
async def get_ffu_log(date: str = Query(...)):
    return JSONResponse(await load_date_form('ffu_log', date))

@app.post('/api/ffu-log/cell')
async def save_ffu_log_cell(request: Request):
    body = await request.json()
    await save_date_form_cell('ffu_log', body['date'], body['key'], body['value'])
    return JSONResponse({'ok': True})

@app.post('/api/ffu-log/bulk')
async def save_ffu_log_bulk(request: Request):
    body = await request.json()
    await save_date_form_bulk('ffu_log', body['date'], body.get('records', []))
    return JSONResponse({'ok': True})

@app.get('/api/boiler-log')
async def get_boiler_log(date: str = Query(...)):
    return JSONResponse(await load_date_form('boiler_log', date))

@app.post('/api/boiler-log/cell')
async def save_boiler_log_cell(request: Request):
    body = await request.json()
    await save_date_form_cell('boiler_log', body['date'], body['key'], body['value'])
    return JSONResponse({'ok': True})

@app.post('/api/boiler-log/bulk')
async def save_boiler_log_bulk(request: Request):
    body = await request.json()
    await save_date_form_bulk('boiler_log', body['date'], body.get('records', []))
    return JSONResponse({'ok': True})

@app.get('/api/safety-log')
async def get_safety_log(date: str = Query(...)):
    return JSONResponse(await load_date_form('safety_log', date))

@app.post('/api/safety-log/cell')
async def save_safety_log_cell(request: Request):
    body = await request.json()
    await save_date_form_cell('safety_log', body['date'], body['key'], body['value'])
    return JSONResponse({'ok': True})

@app.post('/api/safety-log/bulk')
async def save_safety_log_bulk(request: Request):
    body = await request.json()
    await save_date_form_bulk('safety_log', body['date'], body.get('records', []))
    return JSONResponse({'ok': True})


# ── HMI 트렌드 API ────────────────────────────────────────────────────

class _OpenpyxlAdapter:
    """openpyxl 워크시트를 xlrd 호환 인터페이스로 래핑."""
    def __init__(self, ws):
        self._rows = [[c.value for c in row] for row in ws.iter_rows()]

    @property
    def nrows(self): return len(self._rows)

    @property
    def ncols(self): return max((len(r) for r in self._rows), default=0)

    def cell_value(self, r, c):
        try:
            v = self._rows[r][c]
            return '' if v is None else v
        except IndexError:
            return ''


def _read_trend_ws(date: str):
    """날짜 문자열(YYYYMMDD)로 워크시트 반환.
    우선순위: .xls(xlrd) → .xls(xlsx 포맷 BytesIO 우회) → .xlsx(openpyxl)
    확장자가 .xls여도 실제 xlsx 포맷인 파일을 자동으로 처리한다.
    """
    from io import BytesIO as _BIO

    def _load_openpyxl(path_or_bytes):
        # bytes 는 BytesIO 로 감싸서 로드, 경로(str/PathLike)는 그대로 전달
        if isinstance(path_or_bytes, (bytes, bytearray)):
            src = _BIO(bytes(path_or_bytes))
        else:
            src = path_or_bytes
        wb = openpyxl.load_workbook(src, data_only=True)
        ws = wb['Raw Data'] if 'Raw Data' in wb.sheetnames else wb.active
        return _OpenpyxlAdapter(ws)

    for ext in ('.xls', '.xlsx'):
        path = os.path.join(TREND_DIR, f'TREND_{date}{ext}')
        if not os.path.exists(path):
            continue
        try:
            if ext == '.xls':
                # 1) 진짜 xls(BIFF) 시도
                try:
                    wb = xlrd.open_workbook(path)
                    try:
                        return wb.sheet_by_name('Raw Data')
                    except Exception:
                        return wb.sheet_by_index(0)
                except Exception:
                    pass
                # 2) .xls 확장자지만 실제론 xlsx 포맷인 경우 BytesIO 우회
                try:
                    with open(path, 'rb') as f:
                        raw = f.read()
                    return _load_openpyxl(raw)
                except Exception:
                    continue
            else:
                return _load_openpyxl(path)
        except Exception:
            continue
    return None


def _parse_header(ws) -> dict:
    """워크시트에서 헤더 행을 자동 감지하고 컬럼명 → 인덱스 맵을 반환한다.

    탐지 규칙:
    - 처음 6행 안에서 'fab'과 'tag'가 모두 포함된 행을 헤더로 인식
    - 컬럼 헤더 이름은 대소문자·공백을 무시하고 alias 테이블과 매칭
    - 'HH:MM' 패턴의 컬럼은 hour_cols 리스트에 순서대로 수집
    - 헤더를 찾지 못하면 기존 하드코딩 구조로 폴백

    Returns:
        dict with keys: fab, process, location, tag, val,
                        alarm_high, alarm_low, unit,
                        hour_cols (list), data_start_row, _header_row
    """
    # 컬럼 헤더 별칭 테이블 (소문자·공백 제거 후 비교)
    KEY_ALIASES: dict[str, set] = {
        'fab':        {'fab'},
        'process':    {'process', 'proc'},
        'location':   {'location', 'loc'},
        'tag':        {'tag', 'tagname', 'tag_name'},
        'val':        {'val', 'type', 'valtype'},
        'alarm_high': {'highalarm', 'alarm_high', 'high_alarm', 'hialarm',
                       'highalarmvalue', 'alarmhigh'},
        'alarm_low':  {'lowalarm', 'alarm_low', 'low_alarm', 'loalarm',
                       'lowalarmvalue', 'alarmlow'},
        'unit':       {'단위', 'unit', 'engunit', 'eng.unit', 'eng_unit'},
    }

    def _norm(v) -> str:
        return str(v).strip().lower().replace(' ', '').replace('.', '') if v not in ('', None) else ''

    # ncols 안전 취득 (xlrd는 .ncols 속성, _OpenpyxlAdapter도 동일하게 추가됨)
    try:
        ncols = ws.ncols
    except AttributeError:
        ncols = 60

    for r in range(min(ws.nrows, 6)):
        row_raw   = [ws.cell_value(r, c) for c in range(ncols)]
        row_norm  = [_norm(v) for v in row_raw]

        # 헤더 행 판별: 정규화 후 'fab'과 'tag'가 모두 있어야 함
        if 'fab' not in row_norm or 'tag' not in row_norm:
            continue

        col_map: dict = {'_header_row': r, 'data_start_row': r + 1, 'hour_cols': []}

        for c, norm_v in enumerate(row_norm):
            assigned = False
            for key, aliases in KEY_ALIASES.items():
                if norm_v in aliases:
                    col_map.setdefault(key, c)   # 첫 번째 매칭 우선
                    assigned = True
                    break
            # 시간 컬럼: 원본 값이 "HH:MM" 형태 (예: "00:00", "23:00")
            if not assigned:
                raw_v = str(row_raw[c]).strip()
                if len(raw_v) == 5 and raw_v[2] == ':':
                    try:
                        h, m = int(raw_v[:2]), int(raw_v[3:])
                        if 0 <= h <= 23 and m == 0:
                            col_map['hour_cols'].append(c)
                    except ValueError:
                        pass

        col_map['hour_cols'].sort()
        return col_map

    # 헤더 미발견 → 하드코딩 폴백 (기존 구조 유지)
    return {
        '_header_row': 1,
        'data_start_row': 2,
        'fab': 1, 'process': 2, 'location': 3, 'tag': 4, 'val': 5,
        'alarm_high': 30, 'alarm_low': 31, 'unit': 32,
        'hour_cols': list(range(6, 30)),
    }


def _fill_dummy(series: list, lo: float, hi: float) -> list:
    return [
        round(random.uniform(lo, hi), 3) if (v == '' or v is None) else round(float(v), 3)
        for v in series
    ]


def _ws_to_raw_tags(ws, fab: str = '', process: str = '') -> dict:
    """헤더 자동 감지 후 태그 dict 반환.
    fab / process 가 빈 문자열이면 전체 반환.
    컬럼 위치는 _parse_header() 가 동적으로 결정하며,
    헤더를 찾지 못하면 기존 하드코딩 구조로 폴백한다.
    """
    cm        = _parse_header(ws)
    fab_col   = cm.get('fab',        1)
    proc_col  = cm.get('process',    2)
    loc_col   = cm.get('location',   3)
    tag_col   = cm.get('tag',        4)
    val_col   = cm.get('val',        5)
    ha_col    = cm.get('alarm_high', 30)
    la_col    = cm.get('alarm_low',  31)
    unit_col  = cm.get('unit',       32)
    hour_cols = cm.get('hour_cols',  list(range(6, 30)))
    d_start   = cm.get('data_start_row', 2)

    tags: dict = {}
    for r in range(d_start, ws.nrows):
        fab_val  = ws.cell_value(r, fab_col)
        proc_val = ws.cell_value(r, proc_col)
        tag_name = ws.cell_value(r, tag_col)

        if not tag_name or str(tag_name).strip() in ('TAG', ''):
            continue
        if fab     and str(fab_val).strip()  != fab:     continue
        if process and str(proc_val).strip() != process: continue

        tag_key  = str(tag_name).strip()
        location = ws.cell_value(r, loc_col)
        vtype    = str(ws.cell_value(r, val_col)).strip().upper()
        unit     = ws.cell_value(r, unit_col)

        if tag_key not in tags:
            tags[tag_key] = {
                'tag':        tag_key,
                'fab':        str(fab_val).strip(),
                'process':    str(proc_val).strip(),
                'location':   '',
                'unit':       '',
                'alarm_high': None,
                'alarm_low':  None,
                'max': [], 'avg': [], 'min': [],
            }

        rec = tags[tag_key]
        if location not in ('', None): rec['location'] = str(location)
        if unit     not in ('', None): rec['unit']     = str(unit)

        # 알람값: 병합셀이므로 값이 있는 행에서만 갱신
        ha = ws.cell_value(r, ha_col) if ha_col is not None else ''
        la = ws.cell_value(r, la_col) if la_col is not None else ''
        if ha not in ('', None):
            try: rec['alarm_high'] = float(ha)
            except (ValueError, TypeError): pass
        if la not in ('', None):
            try: rec['alarm_low']  = float(la)
            except (ValueError, TypeError): pass

        vals = [ws.cell_value(r, c) for c in hour_cols]
        if   vtype == 'MAX': rec['max'] = vals
        elif vtype == 'AVG': rec['avg'] = vals
        elif vtype == 'MIN': rec['min'] = vals

    return tags


def _fill_tag(rec: dict) -> dict:
    """빈 시계열을 알람 범위 내 더미로 채우고 MAX>=AVG>=MIN 보장."""
    lo = rec['alarm_low']  if rec['alarm_low']  is not None else 0.0
    hi = rec['alarm_high'] if rec['alarm_high'] is not None else lo + 10.0
    max_v = _fill_dummy(rec['max'], (lo + hi) * 0.55, hi  * 0.95)
    min_v = _fill_dummy(rec['min'], lo * 1.05,        (lo + hi) * 0.45)
    avg_v = _fill_dummy(rec['avg'], (lo + hi) * 0.4,  (lo + hi) * 0.6)
    for i in range(len(max_v)):
        mx, mn = max_v[i], min_v[i]
        if mx < mn: max_v[i], min_v[i] = mn, mx; mx, mn = mn, mx
        avg_v[i] = round((mx + mn) / 2 + random.uniform(-0.05, 0.05) * abs(hi - lo), 3)
    rec['max'], rec['avg'], rec['min'] = max_v, avg_v, min_v
    return rec


def _list_trend_dates() -> list:
    """TREND_YYYYMMDD.xls / .xlsx 파일에서 날짜 목록을 중복 없이 반환."""
    seen = set()
    for ext in ('*.xls', '*.xlsx'):
        for f in glob.glob(os.path.join(TREND_DIR, f'TREND_????????{ext[1:]}')):
            d = os.path.basename(f)[6:14]
            seen.add(d)
    return sorted(seen)


@app.get('/api/trend/dates')
async def trend_dates():
    return JSONResponse(_list_trend_dates())


@app.get('/api/trend/filters')
async def trend_filters(date: str = Query('')):
    """최신 파일(또는 지정 날짜)에서 FAB → [PROCESS] 목록.
    컬럼 위치는 _parse_header() 로 동적 감지한다."""
    if not date:
        dates = _list_trend_dates()
        if dates:
            date = dates[-1]
    ws = _read_trend_ws(date)
    if ws is None:
        return JSONResponse({'error': 'not_found'}, status_code=404)

    cm       = _parse_header(ws)
    fab_col  = cm.get('fab',  1)
    proc_col = cm.get('process', 2)
    tag_col  = cm.get('tag',  4)
    d_start  = cm.get('data_start_row', 2)

    fab_map: dict = {}
    for r in range(d_start, ws.nrows):
        f   = ws.cell_value(r, fab_col)
        p   = ws.cell_value(r, proc_col)
        tag = ws.cell_value(r, tag_col)
        if not f or not p or not tag or str(tag).strip() in ('TAG', ''):
            continue
        fab_map.setdefault(str(f).strip(), set()).add(str(p).strip())
    return JSONResponse({f: sorted(p) for f, p in sorted(fab_map.items())})


@app.get('/api/trend/data')
async def trend_data(
    date: str = Query(''),
    fab: str = Query(''),
    process: str = Query(''),
):
    """단일 날짜, 24시간 시계열."""
    ws = _read_trend_ws(date)
    if ws is None:
        return JSONResponse({'error': 'not_found'}, status_code=404)
    tags = _ws_to_raw_tags(ws, fab, process)
    return JSONResponse([_fill_tag(rec) for rec in tags.values()])


@app.get('/api/trend/range')
async def trend_range(
    start_date: str = Query(''),
    end_date: str = Query(''),
    fab: str = Query(''),
    process: str = Query(''),
):
    """날짜 범위, 일별 집계(MAX의 일최대 / AVG의 일평균 / MIN의 일최소)."""
    from datetime import date as D, timedelta
    try:
        start = D(int(start_date[:4]), int(start_date[4:6]), int(start_date[6:8]))
        end   = D(int(end_date[:4]),   int(end_date[4:6]),   int(end_date[6:8]))
    except (ValueError, IndexError):
        return JSONResponse({'error': 'invalid_dates'}, status_code=400)

    valid_dates = []
    agg: dict = {}   # tag_name -> {meta, max:[], avg:[], min:[]}
    cur = start
    while cur <= end:
        d = cur.strftime('%Y%m%d')
        ws = _read_trend_ws(d)
        if ws is not None:
            valid_dates.append(d)
            for name, rec in _ws_to_raw_tags(ws, fab, process).items():
                _fill_tag(rec)
                if name not in agg:
                    agg[name] = {'tag': rec['tag'], 'fab': rec.get('fab', ''), 'process': rec.get('process', ''),
                                 'location': rec.get('location', ''),
                                 'unit': rec['unit'],
                                 'alarm_high': rec['alarm_high'], 'alarm_low': rec['alarm_low'],
                                 'max': [], 'avg': [], 'min': []}
                agg[name]['max'].append(round(max(rec['max']), 3))
                agg[name]['avg'].append(round(sum(rec['avg']) / len(rec['avg']), 3))
                agg[name]['min'].append(round(min(rec['min']), 3))
        cur += timedelta(days=1)

    if not valid_dates:
        return JSONResponse({'error': 'no_data'}, status_code=404)
    return JSONResponse({'dates': valid_dates, 'tags': list(agg.values())})


@app.get('/api/trend/cross')
async def trend_cross(
    start_date: str = Query(''),
    start_hour: int = Query(0),
    end_date:   str = Query(''),
    end_hour:   int = Query(23),
    fab:     str = Query(''),
    process: str = Query(''),
):
    """여러 날짜에 걸친 시간별 원본 데이터 (예: 3/20 00:00 ~ 3/30 18:00).
    labels = ["03/20 00:00", "03/20 01:00", ...]
    tags   = [TagData with max/avg/min length == len(labels)]
    """
    from datetime import date as D, timedelta
    try:
        s_date = D(int(start_date[:4]), int(start_date[4:6]), int(start_date[6:]))
        e_date = D(int(end_date[:4]),   int(end_date[4:6]),   int(end_date[6:]))
    except (ValueError, IndexError):
        return JSONResponse({'error': 'invalid_dates'}, status_code=400)
    if s_date > e_date:
        s_date, e_date = e_date, s_date

    labels: list = []
    agg: dict = {}   # tag_name -> {meta, max:[], avg:[], min:[]}

    cur = s_date
    while cur <= e_date:
        d = cur.strftime('%Y%m%d')
        h_lo = start_hour if cur == s_date else 0
        h_hi = end_hour   if cur == e_date else 23
        n_h  = h_hi - h_lo + 1
        prev_len = len(labels)
        date_label = cur.strftime('%m/%d')

        ws = _read_trend_ws(d)
        found: set = set()

        if ws is not None:
            raw = _ws_to_raw_tags(ws, fab, process)
            for name, rec in raw.items():
                _fill_tag(rec)
                found.add(name)
                if name not in agg:
                    agg[name] = {
                        'tag': rec['tag'], 'fab': rec.get('fab', ''),
                        'process': rec.get('process', ''),
                        'location': rec.get('location', ''),
                        'unit': rec['unit'],
                        'alarm_high': rec['alarm_high'], 'alarm_low': rec['alarm_low'],
                        'max': [None] * prev_len,
                        'avg': [None] * prev_len,
                        'min': [None] * prev_len,
                    }
                agg[name]['max'].extend(rec['max'][h_lo:h_hi + 1])
                agg[name]['avg'].extend(rec['avg'][h_lo:h_hi + 1])
                agg[name]['min'].extend(rec['min'][h_lo:h_hi + 1])

        # 이날 데이터 없는 태그 → None 패딩
        for name in agg:
            if name not in found:
                agg[name]['max'].extend([None] * n_h)
                agg[name]['avg'].extend([None] * n_h)
                agg[name]['min'].extend([None] * n_h)

        for h in range(h_lo, h_hi + 1):
            labels.append(f"{date_label} {h:02d}:00")

        cur += timedelta(days=1)

    if not labels:
        return JSONResponse({'error': 'no_data'}, status_code=404)
    return JSONResponse({'labels': labels, 'tags': list(agg.values())})


# ── Work Orders ──────────────────────────────────────────────────────
@app.post('/api/work-orders/upload')
async def upload_work_orders(file: UploadFile = File(...)):
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

    header_aliases = {
        'wo_no': ['Wo No', 'WO No', 'WO번호', '작업번호'],
        'work_type': ['작업 유형', '작업유형'],
        'work_name': ['작업명'],
        'equip_code': ['설비코드'],
        'equip_name': ['설비명'],
        'equip_type': ['설비종류'],
        'location': ['위치 L4', '위치'],
        'start_date': ['시작일'],
        'end_date': ['종료일'],
        'department': ['작업부서'],
        'worker': ['작업자'],
        'writer': ['작성자'],
        'wo_status': ['WO 상태', 'WO상태', '상태'],
    }

    def normalize_header(value) -> str:
        return str(value or '').strip().replace('\n', ' ')

    def find_header_row(ws):
        for row_no, row in enumerate(ws.iter_rows(values_only=True), start=1):
            headers = [normalize_header(c) for c in row]
            if any(h in header_aliases['wo_no'] for h in headers):
                return row_no, headers
        return None, []

    def build_col_map(headers: list[str]) -> dict:
        result = {}
        for key, aliases in header_aliases.items():
            result[key] = next((i for i, h in enumerate(headers) if h in aliases), None)
        return result

    def cell_to_text(value) -> str:
        if value is None:
            return ''
        if hasattr(value, 'strftime'):
            return value.strftime('%Y-%m-%d')
        return str(value).strip()

    records = []
    seen_wo_no: set[str] = set()
    for ws in wb.worksheets:
        header_row, headers = find_header_row(ws)
        if header_row is None:
            continue

        col_map = build_col_map(headers)
        if col_map.get('wo_no') is None:
            continue

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or all(c is None for c in row):
                continue

            def get(key: str) -> str:
                idx = col_map.get(key)
                if idx is None or idx >= len(row):
                    return ''
                return cell_to_text(row[idx])

            raw_wo_no = get('wo_no')
            if not raw_wo_no or raw_wo_no in {'예시값', '임시값', '임시'}:
                continue

            wo_no = raw_wo_no
            if wo_no in seen_wo_no:
                wo_no = f'{ws.title}-{raw_wo_no}'
            seen_wo_no.add(wo_no)

            record = {key: get(key) for key in header_aliases}
            record['wo_no'] = wo_no
            records.append(record)

    wb.close()

    if not records:
        return JSONResponse({'inserted': 0, 'updated': 0, 'message': '유효한 데이터 없음'})

    result = await upsert_work_orders(records)
    return JSONResponse(result)


@app.get('/api/work-orders')
async def get_work_orders(
    start_date: str = Query(None), end_date: str = Query(None),
    department: str = Query(None), wo_status: str = Query(None),
    work_type: str = Query(None), equip_type: str = Query(None),
    location: str = Query(None), worker: str = Query(None),
    page: int = Query(1), page_size: int = Query(50),
):
    data = await query_work_orders(
        start_date=start_date, end_date=end_date,
        department=department, wo_status=wo_status,
        work_type=work_type, equip_type=equip_type,
        location=location, worker=worker,
        page=page, page_size=page_size,
    )
    return JSONResponse(data)


@app.get('/api/work-orders/kpi')
async def work_order_kpi(
    start_date: str = Query(None),
    end_date: str = Query(None),
):
    data = await get_work_order_kpi(start_date=start_date, end_date=end_date)
    return JSONResponse(data)


@app.get('/api/work-orders/filter-options')
async def work_order_filter_options():
    data = await get_work_order_filter_options()
    return JSONResponse(data)


# ── 전공장 엑셀 파싱 & DB 저장 ───────────────────────────────────────
ZANGONG_SHEET_MAP = {'1공장': 'f1', '2공장': 'f2', '3공장': 'f3', '신공장': 'fnew'}

ZANGONG_FILES = [
    {'util_key': 'gas',      'folder': 'LNG',      'filename': '도시가스사용량_전공장.xlsx', 'format': 'horizontal'},
    {'util_key': 'nitrogen', 'folder': 'NItrogen', 'filename': '질소사용량_전공장.xlsx',    'format': 'vertical'},
    {'util_key': 'argon',    'folder': 'ARGON',    'filename': '아르곤사용량_전공장.xlsx',  'format': 'vertical'},
]


def _parse_zangong_horizontal(wb, util_key: str) -> list:
    """LNG 가로형: 1행=날짜(열), 2행=예산사용량(budget), 3행=실적사용량(value)"""
    records = []
    for sheet_name, factory_key in ZANGONG_SHEET_MAP.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows_data = list(ws.iter_rows(min_row=1, max_row=3, values_only=True))
        if len(rows_data) < 3:
            continue
        date_row = rows_data[0]
        budg_row = rows_data[1]   # 2행 = 예산사용량 = budget
        val_row  = rows_data[2]   # 3행 = 실적사용량 = value
        for i in range(1, len(date_row)):
            d = date_row[i]
            v = val_row[i] if i < len(val_row) else None
            if d is None or not hasattr(d, 'strftime'):
                continue
            # 실적(value)이 없으면 건너뜀
            if v is None:
                continue
            try:
                fv = float(v)
                if fv <= 0:
                    continue
            except (TypeError, ValueError):
                continue
            fb = None
            if i < len(budg_row) and budg_row[i] is not None:
                try:
                    fb = float(budg_row[i])
                    if fb <= 0:
                        fb = None
                except (TypeError, ValueError):
                    pass
            rec = {'util_key': util_key, 'factory': factory_key,
                   'date': d.strftime('%Y-%m-%d'), 'value': fv}
            if fb is not None:
                rec['budget'] = fb
            records.append(rec)
    return records


def _parse_zangong_vertical(wb, util_key: str) -> list:
    """질소/아르곤 세로형: A열=날짜, B열=사용량
    3공장 시트는 B열=3공장(f3), C열=신공장(fnew) 별도 저장"""
    records = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if sheet_name == '3공장':
            # B열 → f3 (P1/P2/기타), C열 → fnew (P3/G3/UT동)
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None or not hasattr(row[0], 'strftime'):
                    continue
                date_str = row[0].strftime('%Y-%m-%d')
                for col_idx, fk in [(1, 'f3'), (2, 'fnew')]:
                    if col_idx >= len(row) or row[col_idx] is None:
                        continue
                    try:
                        fv = float(row[col_idx])
                        if fv > 0:
                            records.append({'util_key': util_key, 'factory': fk,
                                            'date': date_str, 'value': fv})
                    except (TypeError, ValueError):
                        pass
        elif sheet_name in ZANGONG_SHEET_MAP:
            factory_key = ZANGONG_SHEET_MAP[sheet_name]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None or not hasattr(row[0], 'strftime'):
                    continue
                if len(row) <= 1 or row[1] is None:
                    continue
                try:
                    fv = float(row[1])
                    if fv <= 0:
                        continue
                except (TypeError, ValueError):
                    continue
                records.append({'util_key': util_key, 'factory': factory_key,
                                'date': row[0].strftime('%Y-%m-%d'), 'value': fv})
    return records


@app.post('/api/energy/import')
async def import_energy_files():
    all_records = []
    errors = []
    for cfg in ZANGONG_FILES:
        path = os.path.join(PROJECT_ROOT, cfg['folder'], cfg['filename'])
        if not os.path.exists(path):
            errors.append(f"{cfg['filename']} 파일 없음")
            continue
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            if cfg['format'] == 'horizontal':
                records = _parse_zangong_horizontal(wb, cfg['util_key'])
            else:
                records = _parse_zangong_vertical(wb, cfg['util_key'])
            wb.close()
            all_records.extend(records)
        except Exception as e:
            errors.append(f"{cfg['filename']}: {str(e)}")

    result = await upsert_energy_usage(all_records) if all_records else {'count': 0}
    return JSONResponse({**result, 'errors': errors})


@app.get('/api/energy/db')
async def get_energy_from_db():
    return JSONResponse(await query_energy_db())


@app.delete('/api/energy/entries/range')
async def delete_energy_range_endpoint(request: Request):
    global _energy_cache
    body  = await request.json()
    count = await delete_energy_usage_range(
        body['util_key'], body.get('factories', []),
        body['start_date'], body['end_date'],
    )
    _energy_cache = {'data': None, 'ts': 0}
    return JSONResponse({'deleted': count})


# ── 에너지 엑셀 write-back 헬퍼 ───────────────────────────────────────
_REVERSE_SHEET_MAP = {v: k for k, v in ZANGONG_SHEET_MAP.items()}


def _update_horizontal_excel(wb, changes_by_factory: dict) -> list:
    """가로형(gas): row1=날짜열, row2=예산, row3=실적. 없는 날짜는 오류 기록."""
    errors = []
    for factory_key, changes in changes_by_factory.items():
        sheet_name = _REVERSE_SHEET_MAP.get(factory_key)
        if not sheet_name or sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        date_col: dict = {}
        for col in range(2, ws.max_column + 2):
            cell = ws.cell(row=1, column=col)
            if cell.value is None:
                break
            if hasattr(cell.value, 'strftime'):
                date_col[cell.value.strftime('%Y-%m-%d')] = col
        for chg in changes:
            col = date_col.get(chg['date'])
            if col is None:
                errors.append(f"날짜 {chg['date']} 열 없음 ({sheet_name})")
                continue
            if chg['action'] == 'delete':
                ws.cell(row=3, column=col).value = None
                ws.cell(row=2, column=col).value = None
            else:
                ws.cell(row=3, column=col).value = chg.get('value')
                if chg.get('budget') is not None:
                    ws.cell(row=2, column=col).value = chg['budget']
    return errors


def _update_vertical_excel(wb, changes_by_factory: dict) -> list:
    """세로형(nitrogen/argon): colA=날짜, colB=값 (f3/fnew는 '3공장' 시트 B/C열)."""
    from datetime import datetime as _dt
    FACTORY_SHEET_COL = {
        'f1':   ('1공장', 2),
        'f2':   ('2공장', 2),
        'f3':   ('3공장', 2),
        'fnew': ('3공장', 3),
    }
    errors = []
    # 시트별 그룹핑 (f3/fnew가 같은 '3공장' 시트 공유)
    sheet_col_changes: dict = {}
    for factory_key, changes in changes_by_factory.items():
        info = FACTORY_SHEET_COL.get(factory_key)
        if not info:
            continue
        sheet_name, val_col = info
        sheet_col_changes.setdefault(sheet_name, {}).setdefault(val_col, []).extend(changes)

    for sheet_name, col_changes in sheet_col_changes.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        date_row: dict = {}
        for row in range(2, ws.max_row + 2):
            cell = ws.cell(row=row, column=1)
            if cell.value is None:
                break
            if hasattr(cell.value, 'strftime'):
                date_row[cell.value.strftime('%Y-%m-%d')] = row
        for val_col, changes in col_changes.items():
            for chg in changes:
                row = date_row.get(chg['date'])
                if row is None:
                    if chg['action'] != 'delete':
                        new_row = ws.max_row + 1
                        try:
                            d = _dt.strptime(chg['date'], '%Y-%m-%d').date()
                            ws.cell(row=new_row, column=1).value = d
                            ws.cell(row=new_row, column=val_col).value = chg.get('value')
                            date_row[chg['date']] = new_row
                        except Exception as e:
                            errors.append(f"날짜 {chg['date']} 추가 실패: {e}")
                else:
                    if chg['action'] == 'delete':
                        ws.cell(row=row, column=val_col).value = None
                    else:
                        ws.cell(row=row, column=val_col).value = chg.get('value')
    return errors


@app.post('/api/energy/entries/save')
async def save_energy_entries(request: Request):
    """변경사항을 DB + Excel 파일에 동시 저장, 캐시 무효화."""
    global _energy_cache
    body    = await request.json()
    changes = body.get('changes', [])
    excel_errors: list = []

    upserts = [c for c in changes if c.get('action') != 'delete']
    deletes = [c for c in changes if c.get('action') == 'delete']

    saved = 0
    if upserts:
        res = await upsert_energy_usage([{
            'util_key': c['util_key'], 'factory': c['factory'],
            'date': c['date'], 'value': c['value'],
            'budget': c.get('budget'),
        } for c in upserts])
        saved = res.get('count', 0)

    for d in deletes:
        await delete_energy_usage(d['util_key'], d['factory'], d['date'])

    # Excel write-back: util_key → factory_key → [changes]
    by_util: dict = {}
    for c in changes:
        by_util.setdefault(c.get('util_key', ''), {}).setdefault(c.get('factory', ''), []).append(c)

    for cfg in ZANGONG_FILES:
        uk = cfg['util_key']
        if uk not in by_util:
            continue
        path = os.path.join(PROJECT_ROOT, cfg['folder'], cfg['filename'])
        if not os.path.exists(path):
            excel_errors.append(f"{cfg['filename']} 파일 없음")
            continue
        try:
            wb = openpyxl.load_workbook(path)
            errs = (_update_horizontal_excel(wb, by_util[uk])
                    if cfg['format'] == 'horizontal'
                    else _update_vertical_excel(wb, by_util[uk]))
            excel_errors.extend(errs)
            wb.save(path)
            wb.close()
        except Exception as e:
            excel_errors.append(f"{cfg['filename']}: {str(e)}")

    _energy_cache = {'data': None, 'ts': 0}
    return JSONResponse({'saved': saved, 'deleted': len(deletes), 'excel_errors': excel_errors})


# ── 에너지 폴더 스캔 ──────────────────────────────────────────────────
import re

# 프로젝트 루트 (backend 상위 폴더)
PROJECT_ROOT = os.path.normpath(os.path.join(os.path.dirname(__file__), '..'))

UTIL_DIR_MAP: dict[str, str] = {
    'LNG':      'gas',
    'STEAM':    'steam',
    'NItrogen': 'nitrogen',
    'NITROGEN': 'nitrogen',
    'Nitrogen': 'nitrogen',
    'ARGON':    'argon',
    'Argon':    'argon',
}

FACTORY_DIR_MAP: dict[str, str] = {
    '1공장': 'f1',
    '2공장': 'f2',
    '3공장': 'f3',
    '신공장': 'fnew',
}


def _parse_year_month(folder_name: str, file_name: str):
    """폴더명 '26년 4월' 또는 파일명 *20260414* 에서 (year, month) 추출"""
    m = re.search(r'(\d{2,4})년\s*(\d{1,2})월', folder_name)
    if m:
        y = int(m.group(1))
        return (y + 2000 if y < 100 else y), int(m.group(2))
    m = re.search(r'(\d{4})(\d{2})\d{2}', file_name)
    if m:
        return int(m.group(1)), int(m.group(2))
    return None, None


def _parse_energy_excel(path: str, year: int, month: int) -> list[dict]:
    """엑셀 파일에서 일별 사용량 파싱. A=날(1-31), B=사용량"""
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        entries = []
        for row in ws.iter_rows(min_row=5, values_only=True):
            if not row or row[0] is None:
                continue
            try:
                day = int(row[0])
                if not (1 <= day <= 31):
                    continue
                # B열(index 1) 우선, 없으면 C열(index 2)
                val = row[1] if (len(row) > 1 and row[1] is not None) else (
                      row[2] if len(row) > 2 else None)
                if val is None:
                    continue
                v = float(val)
                if v <= 0:
                    continue
                date = f"{year}-{str(month).zfill(2)}-{str(day).zfill(2)}"
                entries.append({'date': date, 'value': v})
            except (ValueError, TypeError):
                continue
        wb.close()
        return entries
    except Exception:
        return []


# 스캔 캐시 (5분 TTL)
import time as _time
_energy_cache: dict = {'data': None, 'ts': 0}
_CACHE_TTL = 300  # seconds


def _scan_energy_dirs() -> dict:
    result: dict = {
        'gas':      {'f1': [], 'f2': [], 'f3': [], 'fnew': []},
        'steam':    {'f1': [], 'f2': [], 'f3': [], 'fnew': []},
        'nitrogen': {'f1': [], 'f2': [], 'f3': [], 'fnew': []},
        'argon':    {'f1': [], 'f2': [], 'f3': [], 'fnew': []},
    }

    for util_dir, util_key in UTIL_DIR_MAP.items():
        util_path = os.path.join(PROJECT_ROOT, util_dir)
        if not os.path.isdir(util_path):
            continue
        for factory_dir, factory_key in FACTORY_DIR_MAP.items():
            factory_path = os.path.join(util_path, factory_dir)
            if not os.path.isdir(factory_path):
                continue
            # 직접 파일 + 하위 월별 폴더 모두 스캔
            for root, _dirs, files in os.walk(factory_path):
                folder_name = os.path.basename(root)
                for fname in sorted(files):
                    if not fname.lower().endswith('.xlsx') or fname.startswith('~'):
                        continue
                    year, month = _parse_year_month(folder_name, fname)
                    if year is None:
                        continue
                    fpath = os.path.join(root, fname)
                    entries = _parse_energy_excel(fpath, year, month)
                    result[util_key][factory_key].extend(entries)

    # 날짜 중복 제거 (같은 날짜면 마지막 값 유지)
    for uk in result:
        for fk in result[uk]:
            date_map: dict[str, float] = {}
            for e in result[uk][fk]:
                date_map[e['date']] = e['value']
            result[uk][fk] = [
                {'date': d, 'value': v}
                for d, v in sorted(date_map.items())
            ]
    return result


@app.get('/api/energy-files')
async def get_energy_files(refresh: bool = Query(False)):
    global _energy_cache
    now = _time.time()
    if not refresh and _energy_cache['data'] and (now - _energy_cache['ts']) < _CACHE_TTL:
        return JSONResponse(_energy_cache['data'])
    data = _scan_energy_dirs()
    _energy_cache = {'data': data, 'ts': now}
    return JSONResponse(data)


# ── Global Exception Handler ─────────────────────────────────────────
from fastapi.exceptions import RequestValidationError
from fastapi.responses import JSONResponse as _JSONResponse

@app.exception_handler(Exception)
async def global_exception_handler(req, exc):
    return _JSONResponse({'error': str(exc)}, status_code=500)


# ── Dashboard Summary ─────────────────────────────────────────────────
@app.get('/api/dashboard/summary')
async def dashboard_summary():
    data = await get_dashboard_summary()
    return JSONResponse(data)


# ── Cost Records ──────────────────────────────────────────────────────
@app.get('/api/cost-records')
async def api_get_cost_records(year: str = Query(...), month: str = Query(...)):
    rows = await get_cost_records(year, month)
    return JSONResponse(rows)

@app.post('/api/cost-records')
async def api_upsert_cost_record(request: Request):
    b = await request.json()
    record_id = await upsert_cost_record(
        b['month'], b['item'], b.get('category', '수선비'),
        int(b.get('budget', 0)), int(b.get('actual', 0)), b.get('owner', '1P')
    )
    return JSONResponse({'ok': True, 'id': record_id})

@app.delete('/api/cost-records/{record_id}')
async def api_delete_cost_record(record_id: int):
    await delete_cost_record(record_id)
    return JSONResponse({'ok': True})


# ── Inventory ─────────────────────────────────────────────────────────
@app.get('/api/inventory/items')
async def api_get_inventory_items():
    return JSONResponse(await get_inventory_items())

@app.post('/api/inventory/items')
async def api_add_inventory_item(request: Request):
    b = await request.json()
    new_id = await add_inventory_item({
        'name': b.get('name', ''), 'spec': b.get('spec', ''),
        'current_qty': int(b.get('current_qty', 0)), 'safety_qty': int(b.get('safety_qty', 0)),
        'price': int(b.get('price', 0)), 'rack': b.get('rack', ''),
        'grade': b.get('grade', 'A'), 'order_point': int(b.get('order_point', 0)),
        'order_qty': int(b.get('order_qty', 0)), 'lead_time': b.get('lead_time', ''),
    })
    return JSONResponse({'ok': True, 'id': new_id})

@app.patch('/api/inventory/items/{item_id}/qty')
async def api_update_inventory_qty(item_id: str, request: Request):
    b = await request.json()
    new_qty = await update_inventory_qty(item_id, int(b.get('delta', 0)))
    return JSONResponse({'ok': True, 'current_qty': new_qty})

@app.get('/api/inventory/issues')
async def api_get_inventory_issues(start: str = Query(None), end: str = Query(None)):
    return JSONResponse(await get_inventory_issues(start, end))

@app.post('/api/inventory/issues')
async def api_add_inventory_issue(request: Request):
    b = await request.json()
    await add_inventory_issue(b['item_id'], int(b['qty']), b.get('user_name', ''), b.get('issued_at', ''))
    return JSONResponse({'ok': True})


# ── Compliance ────────────────────────────────────────────────────────
@app.get('/api/compliance/officers')
async def api_get_officers():
    return JSONResponse(await get_compliance_officers())

@app.post('/api/compliance/officers')
async def api_add_officer(request: Request):
    b = await request.json()
    new_id = await add_compliance_officer(b['role'], b['name'], b.get('dept', ''), b.get('license', ''), b.get('due_date', ''))
    return JSONResponse({'ok': True, 'id': new_id})

@app.delete('/api/compliance/officers/{record_id}')
async def api_delete_officer(record_id: int):
    await delete_compliance_officer(record_id)
    return JSONResponse({'ok': True})

@app.get('/api/compliance/educations')
async def api_get_educations():
    return JSONResponse(await get_compliance_educations())

@app.post('/api/compliance/educations')
async def api_add_education(request: Request):
    b = await request.json()
    new_id = await add_compliance_education(b['name'], b['course'], b.get('done_date', ''), b.get('next_date', ''))
    return JSONResponse({'ok': True, 'id': new_id})

@app.delete('/api/compliance/educations/{record_id}')
async def api_delete_education(record_id: int):
    await delete_compliance_education(record_id)
    return JSONResponse({'ok': True})

@app.get('/api/compliance/inspections')
async def api_get_inspections():
    return JSONResponse(await get_compliance_inspections())

@app.post('/api/compliance/inspections')
async def api_add_inspection(request: Request):
    b = await request.json()
    new_id = await add_compliance_inspection(b['target'], b.get('area', ''), b.get('due_date', ''), b.get('agency', ''))
    return JSONResponse({'ok': True, 'id': new_id})

@app.delete('/api/compliance/inspections/{record_id}')
async def api_delete_inspection(record_id: int):
    await delete_compliance_inspection(record_id)
    return JSONResponse({'ok': True})


# ── Preventive Inspection ─────────────────────────────────────────────
@app.get('/api/preventive/plans')
async def api_get_preventive_plans():
    return JSONResponse(await get_preventive_plans())

@app.post('/api/preventive/plans')
async def api_add_preventive_plan(request: Request):
    b = await request.json()
    new_id = await add_preventive_plan({
        'equipment': b.get('equipment', ''), 'type': b.get('type', '정기점검'),
        'cycle': b.get('cycle', '월 1회'), 'due_date': b.get('due_date', ''),
        'responsible': b.get('responsible', ''), 'plant': b.get('plant', 'P1'),
        'status': b.get('status', '예정'), 'sheet_id': b.get('sheet_id', ''),
    })
    return JSONResponse({'ok': True, 'id': new_id})

@app.patch('/api/preventive/plans/{plan_id}/status')
async def api_update_plan_status(plan_id: str, request: Request):
    b = await request.json()
    await update_preventive_plan_status(plan_id, b['status'])
    return JSONResponse({'ok': True})

@app.get('/api/preventive/results')
async def api_get_preventive_results():
    return JSONResponse(await get_preventive_results())

@app.post('/api/preventive/results')
async def api_add_preventive_result(request: Request):
    b = await request.json()
    new_id = await add_preventive_result({
        'plan_id': b.get('plan_id', ''), 'equipment': b.get('equipment', ''),
        'result_date': b.get('result_date', ''), 'inspector': b.get('inspector', ''),
        'result': b.get('result', '정상'), 'note': b.get('note', ''),
    })
    return JSONResponse({'ok': True, 'id': new_id})

@app.get('/api/preventive/checksheets')
async def api_get_checksheets():
    return JSONResponse(await get_checksheets())

@app.post('/api/preventive/checksheets')
async def api_add_checksheet(request: Request):
    b = await request.json()
    await add_checksheet(b['id'], b['name'], b.get('type', '월간'))
    for item in b.get('items', []):
        await add_checksheet_item(item.get('id', ''), b['id'], item['item'], item.get('method', ''), item.get('standard', ''))
    return JSONResponse({'ok': True})


# ── Alarm Actions ─────────────────────────────────────────────────────
@app.get('/api/alarm-actions')
async def api_get_alarm_actions(start: str = Query(None), end: str = Query(None)):
    return JSONResponse(await get_alarm_actions(start, end))

@app.post('/api/alarm-actions')
async def api_add_alarm_action(request: Request):
    b = await request.json()
    new_id = await add_alarm_action(
        b.get('alarm_id', ''), b.get('equipment', ''), b.get('issue', ''),
        b.get('action_note', ''), b.get('worker', ''), b.get('completed_at', '')
    )
    return JSONResponse({'ok': True, 'id': new_id})


# ── Work Orders — Single Entry & Status Update ────────────────────────
@app.post('/api/work-orders/single')
async def api_upsert_work_order_single(request: Request):
    b = await request.json()
    wo_no = await upsert_work_order_single(b)
    return JSONResponse({'ok': True, 'wo_no': wo_no})

@app.patch('/api/work-orders/{wo_no}/status')
async def api_update_wo_status(wo_no: str, request: Request):
    b = await request.json()
    await update_work_order_status(wo_no, b['status'])
    return JSONResponse({'ok': True})


# ── HMI 트렌드 파일 업로드 ──────────────────────────────────────────────
@app.post('/api/trend/upload')
async def trend_upload(files: List[UploadFile] = File(...)):
    """TREND_YYYYMMDD.xls|xlsx 파일을 HMI_Trend_data 폴더에 저장."""
    os.makedirs(TREND_DIR, exist_ok=True)
    saved, errors = [], []
    for f in files:
        name = f.filename or ''
        m = re.search(r'(\d{8})', name)
        if not m:
            errors.append(f'{name}: 날짜 8자리(YYYYMMDD) 포함 필요')
            continue
        date_str = m.group(1)
        try:
            from datetime import date as _D
            _D(int(date_str[:4]), int(date_str[4:6]), int(date_str[6:]))
        except ValueError:
            errors.append(f'{name}: 유효하지 않은 날짜 {date_str}')
            continue
        ext = os.path.splitext(name)[1].lower()
        if ext not in ('.xls', '.xlsx'):
            ext = '.xls'
        save_path = os.path.join(TREND_DIR, f'TREND_{date_str}{ext}')
        content = await f.read()
        with open(save_path, 'wb') as out:
            out.write(content)
        saved.append(date_str)
    return JSONResponse({'saved': saved, 'errors': errors})


# ── 프론트엔드 정적 파일 서빙 ─────────────────────────────────────────
# Mount static assets (js, css, images) from dist/assets
if os.path.isdir(os.path.join(DIST_DIR, 'assets')):
    app.mount('/ut-mech/assets', StaticFiles(directory=os.path.join(DIST_DIR, 'assets')), name='assets')


@app.get('/{full_path:path}')
async def serve_frontend(full_path: str):
    file_path = os.path.join(DIST_DIR, full_path)
    if full_path and os.path.isfile(file_path):
        return FileResponse(file_path)
    index = os.path.join(DIST_DIR, 'index.html')
    return FileResponse(index)


# ── 엔트리포인트 ──────────────────────────────────────────────────────
if __name__ == '__main__':
    import uvicorn
    _port = int(os.getenv('PORT', 5000))
    _host = os.getenv('HOST', '0.0.0.0')
    print(f"  http://localhost:{_port}/ut-mech/")
    uvicorn.run(app, host=_host, port=_port)
